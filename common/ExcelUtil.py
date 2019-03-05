from openpyxl import load_workbook

class ParseExcel(object):

        def __init__(self,filename,sheetname):
                #获取文件
                self.book = load_workbook(filename)
                #获取sheet
                self.sheet = self.book.get_sheet_by_name(sheetname)
                # self.maxRowNum = self.sheet.max_row


        def getDataFromSheet(self):
               #设置最外层为列表形式
                datalist=[]

                for line in self.sheet:
                        #设置内层为字典形式
                        app = {}
                        locallist = []
                        # print(line[0].value)
                        # print(line[1].value)
                        #设置字典的第一列数据为username
                        app['username'] = line[0].value
                        # 设置字典的第二列数据为password
                        app['password'] = line[1].value
                        #把字典放在列表中
                        datalist.append(app)

                        # tmplist = []
                        # tmplist.append(line[1].value)
                        # tmplist.append(line[2].value)
                        # datalist.append(tmplist)
                   #返回一个列表，里面包含了字典
                return datalist


if __name__ == '__main__':
        filename ="C:\\Users\\HP\\Desktop\\1.xlsx"
        sheetname = "Sheet1"
        pe =ParseExcel(filename,sheetname)
        print(pe.getDataFromSheet())